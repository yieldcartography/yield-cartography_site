#!/usr/bin/env python3
"""
yieldcartography.com — site build script.

Reads CSV inputs from the local YIELDS folder and writes the dashboard JSON
to dist/data/yields.json. The dashboards in dist/curves/, dist/term-premia/,
dist/liquidity/, and dist/eh-tests/ all fetch this single JSON at page-load
via /data/yields.json.

Usage:
    python build.py                                         # writes to ./dist/data/yields.json
    python build.py --out-dir ~/Documents/YC_site/dist      # writes directly into the deploy repo
    python build.py --yields-dir /path/to/csvs              # override CSV source dir
"""
import argparse
import re
import pandas as pd
import numpy as np
import json
import os
import datetime as dt
from pathlib import Path

YIELDS_DIR = Path('/Users/marcinhdec/YIELDS')
SITE_DIR   = Path(__file__).parent
DIST       = SITE_DIR / 'dist'
DATA_DIR   = DIST / 'data'

# Standard tenors used everywhere on the site (matches eh_panel_zero/gsw_us/ecb_ea schema)
TENORS_STD = [0.25, 0.5, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
# Suffixes in the EH zero CSVs
TENOR_COLS = ['y_3m', 'y_6m', 'y_1y', 'y_2y', 'y_3y', 'y_4y', 'y_5y',
              'y_6y', 'y_7y', 'y_8y', 'y_9y', 'y_10y']
TENOR_LABELS = ['3m', '6m', '1y', '2y', '3y', '4y', '5y', '6y', '7y', '8y', '9y', '10y']
SNAP_TENORS_FULL = [0.25, 0.5, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]

# 1y forwards: f_0_1, f_1_2, ..., f_9_10
FWD_COLS = [f'f_{i}_{i+1}' for i in range(10)]

# ACM / BRW available rf-yield tenors
ACM_RF_TENORS = [1, 2, 5, 10]
BRW_RF_TENORS = [1, 2, 3, 5, 7, 10]


# ------------------------------ NSS helpers ------------------------------ #

def nss_y(tau, b0, b1, b2, b3, t1, t2):
    h1 = (1 - np.exp(-tau / t1)) / (tau / t1)
    h2 = h1 - np.exp(-tau / t1)
    h3 = (1 - np.exp(-tau / t2)) / (tau / t2) - np.exp(-tau / t2)
    return b0 + b1 * h1 + b2 * h2 + b3 * h3


def nss_zero_pct(tau, row):
    return nss_y(tau, row.beta0, row.beta1, row.beta2, row.beta3, row.tau1, row.tau2) * 100


def fwd_1y(row, t_start):
    """1-year forward starting at t_start, in percent."""
    z1 = nss_y(t_start,     row.beta0, row.beta1, row.beta2, row.beta3, row.tau1, row.tau2)
    z2 = nss_y(t_start + 1, row.beta0, row.beta1, row.beta2, row.beta3, row.tau1, row.tau2)
    f = z2 * (t_start + 1) - z1 * t_start
    return f * 100


def safe_list(s):
    return [None if pd.isna(x) else float(x) for x in s]


def safe_round(s, n=4):
    return [None if pd.isna(x) else round(float(x), n) for x in s]


def _try_read(path, **kw):
    p = YIELDS_DIR / path
    if not p.exists():
        print(f'  WARN: {path} not found, skipping')
        return None
    return pd.read_csv(p, **kw)


# ------------------------------ data builders ------------------------------ #

def build():
    # Core NSS + ACM/BRW
    nss = pd.read_csv(YIELDS_DIR / 'nss_params_history.csv', parse_dates=['tradedate'])
    acm = pd.read_csv(YIELDS_DIR / 'acm_term_premia_history.csv', parse_dates=['tradedate'])
    brw = pd.read_csv(YIELDS_DIR / 'acm_term_premia_brw_history.csv', parse_dates=['tradedate'])

    eh_pl = _try_read('eh_panel_zero.csv',         parse_dates=['date'])
    eh_us = _try_read('gsw_us.csv',                parse_dates=['date'])
    eh_ea = _try_read('ecb_ea.csv',                parse_dates=['date'])
    fwd   = _try_read('eh_panel_forward.csv',      parse_dates=['date'])
    spf   = _try_read('prof_forecasters_implied.csv', parse_dates=['fcast_date'])
    bond  = _try_read('liquidity_bond_day_full.csv',  parse_dates=['date'])
    # Daily-updated bond panel (BondSpot fixings + Min-Fin weights). Used as a
    # fallback for snapshot dates that are newer than the monthly liquidity
    # rebuild — without it, every snapshot past the last liquidity refresh has
    # an empty bond panel on the site.
    bases = _try_read('bases_merged_all_with_weights_mix.csv', parse_dates=['date'])

    # KDPW / Min-Fin publish monthly turnover with up to a ~2-month lag, so on a
    # fresh snapshot the most recent `turnover_monthly_mln` for a bond can be 0
    # purely because the report isn't out yet. Forward-fill per ISIN on
    # non-zero values so the panel always displays the latest *reported*
    # monthly turnover for each bond — exactly the behaviour the daily slide
    # already uses.
    bond  = _ffill_turnover(bond)
    bases = _ffill_turnover(bases)

    # Min-Fin OUTRIGHT-only monthly turnover, read directly from the xlsx
    # sheet 'Obligacje(T-bonds)_outright' (live) + 'Obligacje(T-bonds)' (static
    # pre-2014, already outright). Used for the display turnover panel —
    # bases.turnover_monthly_mln keeps using the TOTAL (outright+repo+bsb)
    # for backwards-compat with the NSS weight calculation.
    mf_outright = _read_minfin_outright_long(YIELDS_DIR)
    if not mf_outright.empty:
        print(f'  Min-Fin outright: {len(mf_outright):,} (Seria, month) rows, '
              f'{mf_outright["Month"].min()} to {mf_outright["Month"].max()}')
    # Min-Fin repo + sell-buy-back (conditional transactions), for the
    # turnover-tab basis selector. Live file only (begins 2014-04).
    mf_repobsb = _read_minfin_repobsb_long(YIELDS_DIR)
    if not mf_repobsb.empty:
        print(f'  Min-Fin repo+BSB: {len(mf_repobsb):,} (Seria, month) rows, '
              f'{mf_repobsb["Month"].min()} to {mf_repobsb["Month"].max()}')

    # Per-(ISIN, month) BondSpot venue turnover (PLN mln, sum of daily
    # turnover_value) and a (Seria, month) lookup of the Min-Fin outright
    # turnover (PLN bn). Both feed the bond panel's three turnover columns:
    # venue prior month, whole-market outright at the latest reported month,
    # and the venue share of that outright month.
    venue_tov = _venue_monthly_turnover(bases)
    mf_long = {(str(r['Seria']).strip(), r['Month']): float(r['mf_bn'])
               for _, r in mf_outright.iterrows()} if not mf_outright.empty else {}
    mf_months = sorted({m for (_, m) in mf_long.keys()})

    # Index ACM and BRW by trade date for fast lookup per snapshot
    acm_ix = acm.set_index('tradedate').sort_index()
    brw_ix = brw.set_index('tradedate').sort_index()

    # NSS-derived zero rates on standard tenors (in percent)
    for tau in TENORS_STD:
        nss[f'y{tau}'] = nss.apply(lambda r: nss_zero_pct(tau, r), axis=1)

    # ---- weekly time series (Friday close) ----
    nss['week'] = nss.tradedate.dt.to_period('W-FRI')
    nss_w = nss.sort_values('tradedate').groupby('week').last().reset_index(drop=True)
    acm_w = acm.set_index('tradedate').reindex(nss_w['tradedate']).reset_index()
    brw_w = brw.set_index('tradedate').reindex(nss_w['tradedate']).reset_index()

    ts = {
        'dates': nss_w['tradedate'].dt.strftime('%Y-%m-%d').tolist(),
        'y2': safe_list(nss_w['y2']), 'y5': safe_list(nss_w['y5']), 'y10': safe_list(nss_w['y10']),
        'tp10_acm_bp': safe_list(acm_w['tp_10y_bp']), 'tp10_brw_bp': safe_list(brw_w['tp_bc_10y_bp']),
        'tp7_acm_bp':  safe_list(acm_w['tp_7y_bp']),  'tp7_brw_bp':  safe_list(brw_w['tp_bc_7y_bp']),
        'tp5_acm_bp':  safe_list(acm_w['tp_5y_bp']),  'tp5_brw_bp':  safe_list(brw_w['tp_bc_5y_bp']),
        'tp3_acm_bp':  safe_list(acm_w['tp_3y_bp']),  'tp3_brw_bp':  safe_list(brw_w['tp_bc_3y_bp']),
        'tp2_acm_bp':  safe_list(acm_w['tp_2y_bp']),  'tp2_brw_bp':  safe_list(brw_w['tp_bc_2y_bp']),
        'tp1_acm_bp':  safe_list(acm_w['tp_1y_bp']),  'tp1_brw_bp':  safe_list(brw_w['tp_bc_1y_bp']),
        'rf10_acm_pct': safe_list(acm_w['y_rf_10y_pct']), 'rf10_brw_pct': safe_list(brw_w['y_rf_bc_10y_pct']),
        'rf5_acm_pct':  safe_list(acm_w['y_rf_5y_pct']),  'rf5_brw_pct':  safe_list(brw_w['y_rf_bc_5y_pct']),
        'beta0': safe_list(nss_w['beta0'] * 100), 'beta1': safe_list(nss_w['beta1'] * 100),
        'beta2': safe_list(nss_w['beta2'] * 100), 'beta3': safe_list(nss_w['beta3'] * 100),
        'tau1':  safe_list(nss_w['tau1']),        'tau2':  safe_list(nss_w['tau2']),
        'mae_bp': safe_list(nss_w['mae_bp']),     'n_bonds': safe_list(nss_w['n_bonds']),
    }

    # ---- daily companion series for tiles that need day-over-day deltas ----
    # The weekly `ts` above is the canonical series for charts (keeps payload
    # small and is the cadence used by ACM/BRW). For headline tiles where the
    # user expects "since yesterday" semantics on the displayed delta, we
    # publish a separate ts_daily block holding the same zero-rate series at
    # the original daily trading-day frequency.
    nss_d = nss.sort_values('tradedate').reset_index(drop=True)
    ts_daily = {
        'dates': nss_d['tradedate'].dt.strftime('%Y-%m-%d').tolist(),
        'y2':    safe_list(nss_d['y2']),
        'y5':    safe_list(nss_d['y5']),
        'y10':   safe_list(nss_d['y10']),
    }

    # ---- monthly cross-country panel (PL / US / EA, full 12 tenors + spreads) ----
    xc = _build_xcountry(eh_pl, eh_us, eh_ea)

    # ---- cross-country term premia (PL BRW + US ACM + EA ACM, monthly) ----
    xtp = _build_xcountry_tp(YIELDS_DIR / 'gsw_us.csv', YIELDS_DIR / 'ecb_ea.csv', brw)
    xtp_d = _build_xcountry_tp_daily(YIELDS_DIR / 'gsw_us.csv', YIELDS_DIR / 'ecb_ea.csv', brw)
    # Correlation matrices (PL BRW × US ACM, PL BRW × EA ACM)
    corr_pl_us = _tp_xcountry_corr(xtp, 'us')
    corr_pl_ea = _tp_xcountry_corr(xtp, 'ea')

    # ---- PCA loadings on PL zero-rate panel: level / slope / curvature ----
    pca = _build_pca(eh_pl)

    # ---- correlation matrix of TPs (ACM 1/2/5/10 + BRW 1/2/5/10) ----
    corr = _build_corr(ts)

    # ---- liquidity panel (monthly aggregates) ----
    liq = _build_liquidity(bond)

    # ---- EH testing panel (FB1, FB2, CP, OOS-STV, pooled) ----
    eh = _build_eh()

    # ---- monthly snapshots (NSS, bond panel, full-tenor x-country, forwards, SPF) ----
    nss['month'] = nss.tradedate.dt.to_period('M')
    month_ends = nss.sort_values('tradedate').groupby('month').last().reset_index(drop=True)

    snaps = []
    last_panel_isins = set()
    for _, r in month_ends.iterrows():
        d = r['tradedate']
        snap = {
            'date': d.strftime('%Y-%m-%d'),
            'beta0': float(r['beta0']), 'beta1': float(r['beta1']),
            'beta2': float(r['beta2']), 'beta3': float(r['beta3']),
            'tau1':  float(r['tau1']),  'tau2':  float(r['tau2']),
            'n_bonds': int(r['n_bonds']), 'mae_bp': float(r['mae_bp']),
            'nbp': float(r['nbp_rate_act365']) * 100 if pd.notna(r.get('nbp_rate_act365')) else None,
        }

        # NSS zeros at full tenor grid for snapshot
        snap['pl_curve'] = [round(nss_zero_pct(t, r), 4) for t in SNAP_TENORS_FULL]
        snap['pl_tenors'] = SNAP_TENORS_FULL

        # US / EA full-tenor snapshot (nearest prior trading day)
        snap['us_curve'] = _nearest_full(eh_us, d)
        snap['ea_curve'] = _nearest_full(eh_ea, d)

        # Convenience scalars (kept for backward compat with simple charts)
        snap['pl_2y']  = round(nss_zero_pct(2,  r), 4)
        snap['pl_5y']  = round(nss_zero_pct(5,  r), 4)
        snap['pl_10y'] = round(nss_zero_pct(10, r), 4)
        for tag in ('us', 'ea'):
            crv = snap[tag + '_curve']
            snap[f'{tag}_2y']  = crv[3]  if crv else None
            snap[f'{tag}_5y']  = crv[6]  if crv else None
            snap[f'{tag}_10y'] = crv[11] if crv else None

        # Forwards: prefer pre-computed eh_panel_forward.csv if available, fall back to NSS
        snap['pl_fwds'] = _nearest_fwd(fwd, d)
        if snap['pl_fwds'] is None:
            snap['pl_fwds'] = [round(fwd_1y(r, h), 4) for h in range(10)]

        # ACM rf yield curve at integer tenors (expected average short rate over horizon)
        snap['rf_acm'] = _acm_rf(acm_ix, d, ACM_RF_TENORS)
        # BRW rf yield curve (vanilla) and bias-corrected variant
        snap['rf_brw']    = _brw_rf(brw_ix, d, BRW_RF_TENORS, bc=False)
        snap['rf_brw_bc'] = _brw_rf(brw_ix, d, BRW_RF_TENORS, bc=True)
        snap['acm_rf_tenors'] = ACM_RF_TENORS
        snap['brw_rf_tenors'] = BRW_RF_TENORS

        # SPF latest implied path (5 annual readings: y0..y4)
        snap['spf_path'], snap['spf_date'] = _spf_path(spf, d)

        # Bond panel for this snapshot — prefer the liquidity panel
        # (carries pre-computed ttm + segment), fall back to bases for
        # snapshot dates the liquidity panel doesn't cover yet.
        snap['bonds'] = _bond_panel(bond, d, venue_tov, mf_long, mf_months)
        if not snap['bonds']:
            snap['bonds'] = _bond_panel_from_bases(bases, d, venue_tov,
                                                   mf_long, mf_months)
        prior_m, mf_m = _turnover_months(d, mf_months)
        snap['tov_bs_month'] = prior_m
        snap['tov_mf_month'] = mf_m
        last_panel_isins.update(b['isin'] for b in snap['bonds'] if b.get('isin'))

        snaps.append(snap)

    # ---- bond histories: cover ALL ISINs that ever appeared in any monthly snap ----
    all_isins = set()
    for s in snaps:
        for b in s.get('bonds', []):
            if b.get('isin'):
                all_isins.add(b['isin'])
    bond_hist = _bond_hist(bond, all_isins)
    # For ISINs that only appeared in the fallback snapshots (newer than the
    # liquidity rebuild), extend their YTM history from bases so the curves
    # page's "selected bond — YTM history" panel renders the full timeline.
    if bases is not None and not bases.empty:
        bases_ix = bases.set_index('ISIN', drop=False).sort_index()
        for isin in all_isins:
            existing = bond_hist.get(isin)
            try:
                sub = bases_ix.loc[[isin]] if isin in bases_ix.index else None
            except KeyError:
                sub = None
            if sub is None or sub.empty:
                continue
            sub = sub.dropna(subset=['rent_fix_pct']).copy()
            if sub.empty:
                continue
            sub['week'] = pd.to_datetime(sub['date']).dt.to_period('W-FRI')
            sub_w = sub.sort_values('date').groupby('week').last().reset_index(drop=True)
            name = str(sub_w['Nazwa'].iloc[-1]).strip()
            mat = _parse_maturity_from_series(name)
            if mat is None:
                continue
            dates_str = sub_w['date'].dt.strftime('%Y-%m-%d').tolist()
            ytms = [round(float(v), 3) for v in sub_w['rent_fix_pct']]
            ttms = [round(max(0.0, (mat - d).days / 365.25), 4)
                    for d in sub_w['date']]
            if existing:
                # Merge: keep existing dates, append any bases-only dates after the
                # last existing date so we don't double-count overlap.
                last_existing = existing['dates'][-1]
                merge_mask = [d > last_existing for d in dates_str]
                if any(merge_mask):
                    existing['dates'] += [d for d, m in zip(dates_str, merge_mask) if m]
                    existing['ytm']   += [y for y, m in zip(ytms,      merge_mask) if m]
                    existing['ttm']   += [t for t, m in zip(ttms,      merge_mask) if m]
            else:
                bond_hist[isin] = {
                    'name':  name,
                    'dates': dates_str,
                    'ytm':   ytms,
                    'ttm':   ttms,
                }

    return {
        'meta': {
            'first_date': nss_w['tradedate'].iloc[0].strftime('%Y-%m-%d'),
            'last_date':  nss_w['tradedate'].iloc[-1].strftime('%Y-%m-%d'),
            'n_dates_total':       int(len(nss)),
            'n_dates_weekly':      int(len(nss_w)),
            'n_snapshots_monthly': int(len(month_ends)),
            'n_isins_in_history':  int(len(all_isins)),
            'build_ts':            dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
            'schema_version':      'v7',
        },
        'tenor_labels':  TENOR_LABELS,
        'tenor_years':   SNAP_TENORS_FULL,
        'ts':            ts,
        'ts_daily':      ts_daily,
        'xcountry_m':    xc,
        'xcountry_tp_m': xtp,
        'xcountry_tp_d': xtp_d,
        'corr':          corr,
        'corr_pl_us':    corr_pl_us,
        'corr_pl_ea':    corr_pl_ea,
        'pca':           pca,
        'liquidity':     liq,
        'turnover_panel': _build_turnover_panel(bases, mf_outright, mf_repobsb),
        'freshness':     _build_freshness(bases, nss, acm, brw),
        'eh':            eh,
        'snaps':         snaps,
        'bond_hist':     bond_hist,
    }


# ------------------------------ helpers ------------------------------ #

def _build_xcountry(eh_pl, eh_us, eh_ea):
    """Monthly cross-country zero panel — full 12 tenors per country + level/slope/curve for PL."""
    out = {'dates': []}
    for tag in ('pl', 'us', 'ea'):
        for tnr in TENOR_LABELS:
            out[f'{tag}_{tnr}'] = []
        out[f'{tag}_spread'] = []  # 10y - 2y
    if eh_pl is None or eh_us is None or eh_ea is None:
        return out

    for src in (eh_pl, eh_us, eh_ea):
        src['m'] = src['date'].dt.to_period('M')

    pl_m = eh_pl.sort_values('date').groupby('m').last().reset_index()
    us_m = eh_us.sort_values('date').groupby('m').last().reset_index()
    ea_m = eh_ea.sort_values('date').groupby('m').last().reset_index()

    months = sorted(set(pl_m['m']) & set(us_m['m']) & set(ea_m['m']))
    for m in months:
        out['dates'].append(str(m))
        for tag, df in (('pl', pl_m), ('us', us_m), ('ea', ea_m)):
            row = df[df['m'] == m].iloc[0]
            for col, lbl in zip(TENOR_COLS, TENOR_LABELS):
                v = float(row[col]) * 100 if col in row and pd.notna(row[col]) else None
                out[f'{tag}_{lbl}'].append(round(v, 4) if v is not None else None)
            y2 = out[f'{tag}_2y'][-1]; y10 = out[f'{tag}_10y'][-1]
            out[f'{tag}_spread'].append(None if (y2 is None or y10 is None) else round(y10 - y2, 4))
    return out


def _build_pca(eh_pl):
    """Time series of level / slope / curvature factors as already computed in eh_panel_zero."""
    if eh_pl is None or 'level' not in eh_pl.columns:
        return None
    df = eh_pl.dropna(subset=['level']).sort_values('date').copy()
    df['m'] = df['date'].dt.to_period('M')
    df_m = df.groupby('m').last().reset_index()
    return {
        'dates':     df_m['date'].dt.strftime('%Y-%m-%d').tolist(),
        'level':     safe_round(df_m['level']),
        'slope':     safe_round(df_m['slope']),
        'curvature': safe_round(df_m['curvature']),
    }


def _build_corr(ts):
    """Correlation matrix of TP series across horizons (ACM + BRW), in percent return space."""
    keys = ['tp1_acm_bp','tp2_acm_bp','tp5_acm_bp','tp10_acm_bp',
            'tp1_brw_bp','tp2_brw_bp','tp5_brw_bp','tp10_brw_bp']
    labels = ['1y ACM','2y ACM','5y ACM','10y ACM','1y BRW','2y BRW','5y BRW','10y BRW']
    df = pd.DataFrame({k: ts[k] for k in keys}).dropna()
    if df.empty:
        return None
    M = df.corr().round(3)
    return {'labels': labels, 'matrix': M.values.tolist()}


def _build_liquidity(bond):
    if bond is None or bond.empty:
        return None
    cols = ['date','ISIN','bas_bp','ztd','amihud','roll','gamma','corwin_schultz','composite_z']
    cols = [c for c in cols if c in bond.columns]
    df = bond[cols].copy()
    df['m'] = df['date'].dt.to_period('M')
    agg = df.groupby('m').agg(
        n_bonds=('ISIN','nunique'),
        bas_bp_med=('bas_bp','median'),
        bas_bp_p25=('bas_bp', lambda s: s.quantile(0.25)),
        bas_bp_p75=('bas_bp', lambda s: s.quantile(0.75)),
        ztd_mean=('ztd','mean'),
        amihud_med=('amihud','median'),
        roll_med=('roll','median'),
        gamma_med=('gamma','median'),
        cs_med=('corwin_schultz','median'),
        composite_z_mean=('composite_z','mean'),
    ).reset_index()
    agg['date'] = agg['m'].astype(str)
    return {
        'dates':            agg['date'].tolist(),
        'n_bonds':          [int(x) for x in agg['n_bonds']],
        'bas_bp_med':       safe_round(agg['bas_bp_med'], 3),
        'bas_bp_p25':       safe_round(agg['bas_bp_p25'], 3),
        'bas_bp_p75':       safe_round(agg['bas_bp_p75'], 3),
        'ztd_pct':          safe_round(agg['ztd_mean'] * 100, 2),
        'amihud_med':       safe_round(agg['amihud_med'], 6),
        'roll_med':         safe_round(agg['roll_med'], 6),
        'gamma_med':        safe_round(agg['gamma_med'], 8),
        'cs_med':           safe_round(agg['cs_med'], 6),
        'composite_z':      safe_round(agg['composite_z_mean'], 3),
    }


def _build_eh():
    """Bundle EH-test outputs into a JSON-friendly structure."""
    out = {}
    # FB1 / FB2 — main results (PL)
    for tag, fn in (('fb1', 'eh_fb1_results.csv'), ('fb2', 'eh_fb2_results.csv')):
        df = _try_read(fn)
        if df is None:
            continue
        out[tag] = df.fillna('').to_dict(orient='list')

    # FB1 / FB2 — heatmaps (h_m vs n_m)
    for tag, fn in (('fb1_heatmap', 'eh_fb1_heatmap.csv'), ('fb2_heatmap', 'eh_fb2_heatmap.csv')):
        df = _try_read(fn)
        if df is None:
            continue
        # First col is index (h_m), other cols are n_m
        df = df.rename(columns={df.columns[0]: 'h_m'})
        out[tag] = {
            'h_m':    df['h_m'].tolist(),
            'n_m':    [int(c) for c in df.columns[1:]],
            'matrix': df.iloc[:, 1:].values.tolist(),
        }

    # Cross-market FB1 β heatmap (PL / US / EA) on common (h, n) grid — Figure 2 of EH paper
    out['fb1_xm_heatmap'] = _eh_fb1_xm_heatmap()

    # Macro-spanning Figure 5: median R² by horizon for {fb-only, macro-only, joint}
    out['macro_r2_by_h'] = _eh_macro_r2_by_h()

    # CP — second stage (n_y, b_n, etc.)
    cp_main = _try_read('eh_cp_second_stage.csv')
    if cp_main is not None:
        out['cp_second'] = cp_main.fillna('').to_dict(orient='list')

    # CP — multi (single + seagull joint)
    cp_multi = _try_read('eh_cp_multi.csv')
    if cp_multi is not None:
        out['cp_multi'] = cp_multi.fillna('').to_dict(orient='list')

    # OOS-STV (PL, US, EA)
    for tag, fn in (('oos_pl', 'eh_oos_stv_results.csv'),
                    ('oos_us', 'eh_oos_stv_us_results.csv'),
                    ('oos_ea', 'eh_oos_stv_ea_results.csv')):
        df = _try_read(fn)
        if df is None:
            continue
        out[tag] = df.fillna('').to_dict(orient='list')

    # Pooled cross-market FB1
    pooled = _try_read('eh_pooled_results.csv')
    if pooled is not None:
        out['pooled'] = pooled.fillna('').to_dict(orient='list')

    # Macro-spanning regressions
    macro = _try_read('eh_macro_spanning_results.csv')
    if macro is not None:
        out['macro_spanning'] = macro.fillna('').to_dict(orient='list')

    # Descriptive stats
    desc = _try_read('eh_descstat_xmarket.csv')
    if desc is not None:
        out['desc'] = desc.fillna('').to_dict(orient='list')

    # Cross-market FB1 (PL / US / EA)
    for tag, fn in (('fb1_pl', 'eh_fb1_pl_xm.csv'),
                    ('fb1_us', 'eh_fb1_us_xm.csv'),
                    ('fb1_ea', 'eh_fb1_ea_xm.csv')):
        df = _try_read(fn)
        if df is None:
            continue
        out[tag] = df.fillna('').to_dict(orient='list')

    return out


def _nearest_full(df, d):
    if df is None or df.empty:
        return None
    sub = df[df['date'] <= d]
    if sub.empty:
        return None
    row = sub.iloc[-1]
    return [round(float(row[c]) * 100, 4) if c in row and pd.notna(row[c]) else None for c in TENOR_COLS]


def _nearest_fwd(df, d):
    if df is None or df.empty:
        return None
    sub = df[df['date'] <= d]
    if sub.empty:
        return None
    row = sub.iloc[-1]
    return [round(float(row[c]) * 100, 4) if c in row and pd.notna(row[c]) else None for c in FWD_COLS]


def _acm_rf(acm_ix, d, tenors):
    """Pull ACM expected-rate yield curve at the given trade date and tenor list (in percent)."""
    sub = acm_ix.loc[acm_ix.index <= d]
    if sub.empty:
        return None
    row = sub.iloc[-1]
    return [round(float(row[f'y_rf_{t}y_pct']), 4) if f'y_rf_{t}y_pct' in row and pd.notna(row[f'y_rf_{t}y_pct']) else None for t in tenors]


def _brw_rf(brw_ix, d, tenors, bc=False):
    """Pull BRW expected-rate yield curve (vanilla or bias-corrected)."""
    sub = brw_ix.loc[brw_ix.index <= d]
    if sub.empty:
        return None
    row = sub.iloc[-1]
    pre = 'y_rf_bc_' if bc else 'y_rf_'
    return [round(float(row[f'{pre}{t}y_pct']), 4) if f'{pre}{t}y_pct' in row and pd.notna(row[f'{pre}{t}y_pct']) else None for t in tenors]


def _spf_path(spf, d):
    if spf is None or spf.empty:
        return None, None
    sub = spf[spf['fcast_date'] <= d].sort_values('fcast_date')
    if sub.empty:
        return None, None
    row = sub.iloc[-1]
    path = [float(row[c]) if pd.notna(row.get(c)) else None for c in ('y0','y1','y2','y3','y4')]
    return path, row['fcast_date'].strftime('%Y-%m-%d')


def _read_minfin_outright_long(yields_dir):
    """Read the Min-Fin secondary-market turnover, outright-only, as a long
    DataFrame with columns: Seria, Month (Period[M]), mf_bn (PLN bn).

    Reads from two sources and concatenates with live winning on overlap:
      - Live  : Transakcje_po_seriach<NN>.xlsx, sheet 'Obligacje(T-bonds)_outright'
      - Static: Transakcje_pre2014_static.xlsx, sheet 'Obligacje(T-bonds)'
                (already outright-only — the PhD KDPW MD extraction excluded
                repo and BSB).
    Without this filter, the live file's default sheet aggregates outright +
    repo + buy/sell-back, which inflates post-2014-04 totals ~4× and causes
    an artificial jump at the seam between the two source files.
    """
    import warnings as _w
    from pathlib import Path as _P

    yields_dir = _P(yields_dir)

    def _parse_long(path, sheet):
        try:
            with _w.catch_warnings():
                _w.filterwarnings('ignore')
                raw = pd.read_excel(path, sheet_name=sheet, engine='openpyxl',
                                    dtype=object, header=None)
        except Exception as e:
            print(f'  WARN: could not read {path.name} sheet "{sheet}": {e}')
            return pd.DataFrame(columns=['Seria', 'Month', 'mf_bn'])
        date_row = raw.iloc[2, 1:]
        date_vals = [v for v in date_row if v is not None and str(v) != 'nan']
        n_dates = len(date_vals)
        months = pd.PeriodIndex([pd.Timestamp(d).to_period('M') for d in date_vals], freq='M')
        data = raw.iloc[3:, :]
        data = data[data.iloc[:, 0].astype(str).str.strip() != 'Razem/Total']
        data = data[~data.iloc[:, 0].isna()]
        recs = []
        for _, r in data.iterrows():
            seria = str(r.iloc[0]).strip()
            vals = r.iloc[1: 1 + n_dates].values
            for m, v in zip(months, vals):
                try:
                    t = float(v) if v is not None and str(v) != 'nan' else None
                except (TypeError, ValueError):
                    t = None
                if t is not None:
                    recs.append({'Seria': seria, 'Month': m, 'mf_bn': t / 1000.0})
        return pd.DataFrame(recs)

    # Live: pick the freshest Transakcje_po_seriach*.xlsx
    live_candidates = sorted(yields_dir.glob('Transakcje_po_seriach*.xlsx'))
    live_df = pd.DataFrame()
    if live_candidates:
        live_df = _parse_long(live_candidates[-1], 'Obligacje(T-bonds)_outright')

    # Static pre-2014
    static_path = yields_dir / 'Transakcje_pre2014_static.xlsx'
    pre_df = pd.DataFrame()
    if static_path.exists():
        pre_df = _parse_long(static_path, 'Obligacje(T-bonds)')

    if live_df.empty and pre_df.empty:
        return pd.DataFrame(columns=['Seria', 'Month', 'mf_bn'])

    if not live_df.empty and not pre_df.empty:
        live_keys = set(zip(live_df['Seria'], live_df['Month']))
        pre_df = pre_df[~pre_df.apply(lambda r: (r['Seria'], r['Month']) in live_keys, axis=1)]

    out = pd.concat([pre_df, live_df], ignore_index=True)
    return out.sort_values(['Seria', 'Month']).reset_index(drop=True)


def _read_minfin_repobsb_long(yields_dir):
    """Min-Fin repo + sell-buy-back (BSB) monthly turnover, long form
    (Seria, Month[M], mf_bn in PLN bn), summed across the two conditional
    transaction types. Live file only: repo and BSB are not in the pre-2014
    static outright extract, so this series begins 2014-04."""
    import warnings as _w
    from pathlib import Path as _P
    yields_dir = _P(yields_dir)

    def _parse_long(path, sheet):
        try:
            with _w.catch_warnings():
                _w.filterwarnings('ignore')
                raw = pd.read_excel(path, sheet_name=sheet, engine='openpyxl',
                                    dtype=object, header=None)
        except Exception as e:
            print(f'  WARN: could not read {path.name} sheet "{sheet}": {e}')
            return pd.DataFrame(columns=['Seria', 'Month', 'mf_bn'])
        date_row = raw.iloc[2, 1:]
        date_vals = [v for v in date_row if v is not None and str(v) != 'nan']
        n_dates = len(date_vals)
        months = pd.PeriodIndex([pd.Timestamp(d).to_period('M') for d in date_vals], freq='M')
        data = raw.iloc[3:, :]
        data = data[data.iloc[:, 0].astype(str).str.strip() != 'Razem/Total']
        data = data[~data.iloc[:, 0].isna()]
        recs = []
        for _, r in data.iterrows():
            seria = str(r.iloc[0]).strip()
            vals = r.iloc[1: 1 + n_dates].values
            for m, v in zip(months, vals):
                try:
                    t = float(v) if v is not None and str(v) != 'nan' else None
                except (TypeError, ValueError):
                    t = None
                if t is not None:
                    recs.append({'Seria': seria, 'Month': m, 'mf_bn': t / 1000.0})
        return pd.DataFrame(recs)

    live = sorted(yields_dir.glob('Transakcje_po_seriach*.xlsx'))
    if not live:
        return pd.DataFrame(columns=['Seria', 'Month', 'mf_bn'])
    repo = _parse_long(live[-1], 'Obligacje(T-bonds)_repo')
    bsb = _parse_long(live[-1], 'Obligacje(T-bonds)_bsb')
    both = pd.concat([repo, bsb], ignore_index=True)
    if both.empty:
        return both
    both = both.groupby(['Seria', 'Month'], as_index=False)['mf_bn'].sum()
    return both.sort_values(['Seria', 'Month']).reset_index(drop=True)


def _build_turnover_panel(bases, mf_outright=None, mf_repobsb=None):
    """Monthly turnover panel for the liquidity page.

    Three views, all in PLN bn:
      1) Min-Fin total secondary-market turnover stacked by segment
         (short ≤1.5y, belly 1.5–7y, long >7y).
      2) BondSpot venue turnover stacked by segment.
      3) BondSpot share of Min-Fin total = bspot.total / minfin.total · 100.
    Plus a per-bond series (months + minfin_bn + bspot_bn) keyed by ISIN
    for the bond picker.
    """
    if bases is None or bases.empty:
        return None

    b = bases.copy()
    if 'date' not in b.columns or 'ISIN' not in b.columns:
        return None
    b['date'] = pd.to_datetime(b['date'], errors='coerce')
    b = b.dropna(subset=['date'])
    if b.empty:
        return None
    b['month'] = b['date'].dt.to_period('M')

    # Maturity map for segment assignment (changes over time as bonds age)
    series_names = b['Nazwa'].dropna().astype(str).unique()
    mat_map = {n: _parse_maturity_from_series(n) for n in series_names}
    mat_map = {k: v for k, v in mat_map.items() if v is not None}

    def _segment_at(name, month_period):
        mat = mat_map.get(name)
        if mat is None:
            return None
        ms = month_period.to_timestamp()
        if mat <= ms:
            return None
        ttm = (mat - ms).days / 365.25
        if ttm <= 1.5: return 'short'
        if ttm <= 7.0: return 'belly'
        return 'long'

    # -- Min-Fin OUTRIGHT-only monthly turnover, keyed on Seria --
    # We pass `mf_outright` from build() (read directly from the live and static
    # xlsx files in their _outright sheet); fall back to bases.turnover_monthly_mln
    # only if the outright source isn't available, with a loud warning since
    # that fallback inflates post-2014-04 figures ~4× by including repo+bsb.
    if mf_outright is not None and not mf_outright.empty:
        # Need ISIN per Seria → use bases as the ISIN ↔ Seria map.
        seria_to_isin = (b.dropna(subset=['Nazwa', 'ISIN'])
                          .drop_duplicates('Nazwa')
                          .set_index('Nazwa')['ISIN'].to_dict())
        mf = mf_outright.copy()
        mf['ISIN'] = mf['Seria'].map(seria_to_isin).fillna(mf['Seria'])
        mf['Nazwa'] = mf['Seria']
        mf['ref_month_p'] = mf['Month']
        mf = mf.dropna(subset=['mf_bn'])
        # mf already in PLN bn
    elif 'turnover_monthly_mln' in b.columns and 'ref_month' in b.columns:
        print('  WARN: Min-Fin outright sheet not available — falling back to '
              'TOTAL (outright+repo+bsb) from bases. Display totals will be '
              '~4× larger post-2014-04 than outright alone.')
        mf = b.dropna(subset=['ref_month']).copy()
        mf['ref_month_p'] = pd.PeriodIndex(mf['ref_month'].astype(str).str.slice(0, 7), freq='M')
        mf = mf.drop_duplicates(subset=['ISIN', 'ref_month_p']).copy()
        mf['mf_bn'] = pd.to_numeric(mf['turnover_monthly_mln'], errors='coerce') / 1000.0
        mf = mf.dropna(subset=['mf_bn'])
    else:
        mf = pd.DataFrame(columns=['ISIN', 'Nazwa', 'ref_month_p', 'mf_bn'])

    # -- BondSpot monthly turnover: sum of daily turnover_value per ISIN per month --
    # turnover_value in bases is in PLN millions (median ~55, p99 ~610, max ~3400),
    # so divide by 1000 to get PLN bn — NOT by 1e9 which would treat it as raw PLN.
    if 'turnover_value' in b.columns:
        b['tv'] = pd.to_numeric(b['turnover_value'], errors='coerce').fillna(0)
        bs = b.groupby(['ISIN', 'month'], as_index=False).agg(
            bs_mln=('tv', 'sum'),
            name=('Nazwa', 'last'),
        )
        bs['bs_bn'] = bs['bs_mln'] / 1e3   # PLN mln → PLN bn
        bs = bs[bs['bs_bn'] > 0]
    else:
        bs = pd.DataFrame(columns=['ISIN', 'month', 'bs_bn', 'name'])

    # Common monthly index across both sources
    months = sorted(set(mf['ref_month_p'].tolist() if not mf.empty else []) |
                    set(bs['month'].tolist()        if not bs.empty else []))
    months_p = pd.PeriodIndex(months, freq='M')
    month_strs = [str(m) for m in months_p]

    def _segment_totals(df, value_col, name_col, month_col):
        out = {'short': [0.0]*len(months_p), 'belly': [0.0]*len(months_p), 'long': [0.0]*len(months_p)}
        if df is None or df.empty:
            return out
        ix = {m: i for i, m in enumerate(months_p)}
        for _, r in df.iterrows():
            seg = _segment_at(str(r[name_col]), r[month_col])
            if seg is None: continue
            v = float(r[value_col]) if pd.notna(r[value_col]) else 0.0
            i = ix.get(r[month_col])
            if i is None: continue
            out[seg][i] += v
        return {k: [round(x, 4) for x in v] for k, v in out.items()}

    mf_seg = _segment_totals(mf, 'mf_bn', 'Nazwa', 'ref_month_p')
    bs_seg = _segment_totals(bs, 'bs_bn', 'name',  'month')

    # Min-Fin repo + sell-buy-back, mapped to segments on the same monthly grid.
    if mf_repobsb is not None and not mf_repobsb.empty:
        mr = mf_repobsb.copy()
        mr['Nazwa'] = mr['Seria']
        mr['ref_month_p'] = mr['Month']
        mr = mr.dropna(subset=['mf_bn'])
    else:
        mr = pd.DataFrame(columns=['Nazwa', 'ref_month_p', 'mf_bn'])
    rb_seg = _segment_totals(mr, 'mf_bn', 'Nazwa', 'ref_month_p')

    mf_total = [round(mf_seg['short'][i] + mf_seg['belly'][i] + mf_seg['long'][i], 4) for i in range(len(months_p))]
    bs_total = [round(bs_seg['short'][i] + bs_seg['belly'][i] + bs_seg['long'][i], 4) for i in range(len(months_p))]

    # Trim trailing months that have zero turnover from both sources — those are
    # months where Min-Fin hasn't published yet and BondSpot hasn't aggregated
    # the last partial month. Showing them produces an artificial drop on the
    # right edge of the chart.
    last_meaningful = -1
    for i in range(len(months_p) - 1, -1, -1):
        if (mf_total[i] or 0) > 0 or (bs_total[i] or 0) > 0:
            last_meaningful = i
            break
    if last_meaningful >= 0 and last_meaningful < len(months_p) - 1:
        cut = last_meaningful + 1
        months_p = months_p[:cut]
        month_strs = month_strs[:cut]
        for k in mf_seg: mf_seg[k] = mf_seg[k][:cut]
        for k in bs_seg: bs_seg[k] = bs_seg[k][:cut]
        for k in rb_seg: rb_seg[k] = rb_seg[k][:cut]
        mf_total = mf_total[:cut]
        bs_total = bs_total[:cut]

    bs_share = []
    for bsv, mfv in zip(bs_total, mf_total):
        if mfv > 0:
            bs_share.append(round(100.0 * bsv / mfv, 2))
        else:
            bs_share.append(None)

    # Per-bond monthly series (cap history depth for size; only include currently-live bonds)
    today = pd.Timestamp.now().normalize()
    per_bond = {}
    isins = sorted(set(mf['ISIN'].dropna().astype(str).tolist() if not mf.empty else []) |
                   set(bs['ISIN'].dropna().astype(str).tolist() if not bs.empty else []))
    for isin in isins:
        mf_i = mf[mf['ISIN'].astype(str) == isin] if not mf.empty else pd.DataFrame()
        bs_i = bs[bs['ISIN'].astype(str) == isin] if not bs.empty else pd.DataFrame()
        name = ''
        if not mf_i.empty:    name = str(mf_i['Nazwa'].iloc[0])
        elif not bs_i.empty:  name = str(bs_i['name'].iloc[0])
        # only include bonds with at least one non-zero observation
        if (mf_i.empty or mf_i['mf_bn'].fillna(0).sum() <= 0) and \
           (bs_i.empty or bs_i['bs_bn'].fillna(0).sum() <= 0):
            continue
        bm = sorted(set(mf_i['ref_month_p'].tolist() if not mf_i.empty else []) |
                    set(bs_i['month'].tolist()        if not bs_i.empty else []))
        bm_p = pd.PeriodIndex(bm, freq='M')
        mf_s = (mf_i.set_index('ref_month_p')['mf_bn'].reindex(bm_p)
                if not mf_i.empty else pd.Series(index=bm_p, dtype=float))
        bs_s = (bs_i.set_index('month')['bs_bn'].reindex(bm_p)
                if not bs_i.empty else pd.Series(index=bm_p, dtype=float))
        mat = mat_map.get(name)
        per_bond[isin] = {
            'name':     name,
            'isin':     isin,
            'months':   [str(m) for m in bm_p],
            'minfin_bn': [round(float(v), 4) if pd.notna(v) else None for v in mf_s],
            'bspot_bn':  [round(float(v), 4) if pd.notna(v) else None for v in bs_s],
            'live':     bool(mat is not None and mat > today),
        }

    # ---- Outstanding stock by segment (independent of turnover trimming) ----
    out_panel = _build_outstanding_panel(b, mat_map)

    # Min-Fin segment turnover under the three transaction-type bases. The
    # left-hand chart switches between these; outright stays the default and
    # also feeds the BondSpot-share figure.
    _n = len(month_strs)
    def _seg_obj(seg):
        sh, be, lo = seg['short'], seg['belly'], seg['long']
        tot = [round(sh[i] + be[i] + lo[i], 4) for i in range(_n)]
        return {'short_bn': sh, 'belly_bn': be, 'long_bn': lo, 'total_bn': tot}
    all_seg = {k: [round(mf_seg[k][i] + rb_seg[k][i], 4) for i in range(_n)]
               for k in ('short', 'belly', 'long')}
    minfin_basis = {
        'outright': _seg_obj(mf_seg),
        'repobsb':  _seg_obj(rb_seg),
        'all':      _seg_obj(all_seg),
    }

    return {
        'months':            month_strs,
        'minfin_basis':      minfin_basis,
        'minfin_short_bn':   mf_seg['short'],
        'minfin_belly_bn':   mf_seg['belly'],
        'minfin_long_bn':    mf_seg['long'],
        'minfin_total_bn':   mf_total,
        'bspot_short_bn':    bs_seg['short'],
        'bspot_belly_bn':    bs_seg['belly'],
        'bspot_long_bn':     bs_seg['long'],
        'bspot_total_bn':    bs_total,
        'bspot_share_pct':   bs_share,
        'per_bond':          per_bond,
        'outstanding':       out_panel,
    }


def _build_outstanding_panel(b, mat_map):
    """Monthly outstanding stock by segment (short/belly/long), in PLN bn.

    Methodology: take the latest `outstanding_mln` per (ISIN, month) from the
    daily bases panel — outstanding amounts change only at auctions and
    redemptions, so per-bond mid-month variation is essentially zero and the
    last-of-month value is a clean snapshot. Bucket by month-start TTM with
    the same boundaries as the turnover chart.

    Trims trailing months where the total stock collapses to zero or below
    the long-run minimum (data gap, not a real redemption wave).
    """
    if b is None or b.empty or 'outstanding_mln' not in b.columns:
        return None

    b = b.copy()
    b['out_bn'] = pd.to_numeric(b['outstanding_mln'], errors='coerce') / 1000.0
    b = b.dropna(subset=['out_bn'])
    if b.empty:
        return None
    if 'month' not in b.columns:
        b['month'] = pd.to_datetime(b['date'], errors='coerce').dt.to_period('M')
    b = b.dropna(subset=['month'])
    # last value per (ISIN, month)
    snap = (b.sort_values('date')
              .groupby(['ISIN', 'month'], as_index=False)
              .agg(out_bn=('out_bn', 'last'),
                   name=('Nazwa', 'last')))
    snap = snap[snap['out_bn'] > 0]
    if snap.empty:
        return None

    months = sorted(snap['month'].unique())
    months_p = pd.PeriodIndex(months, freq='M')
    out_short = [0.0] * len(months_p)
    out_belly = [0.0] * len(months_p)
    out_long  = [0.0] * len(months_p)
    ix = {m: i for i, m in enumerate(months_p)}

    for _, r in snap.iterrows():
        name = str(r['name']) if pd.notna(r['name']) else ''
        mat = mat_map.get(name)
        if mat is None:
            continue
        ms = r['month'].to_timestamp()
        if mat <= ms:
            continue
        ttm = (mat - ms).days / 365.25
        i = ix.get(r['month'])
        if i is None:
            continue
        v = float(r['out_bn'])
        if ttm <= 1.5:
            out_short[i] += v
        elif ttm <= 7.0:
            out_belly[i] += v
        else:
            out_long[i]  += v

    totals = [round(out_short[i] + out_belly[i] + out_long[i], 4)
              for i in range(len(months_p))]

    # Trim trailing months with zero total (no data yet)
    last_ok = -1
    for i in range(len(months_p) - 1, -1, -1):
        if totals[i] > 0:
            last_ok = i
            break
    if last_ok >= 0 and last_ok < len(months_p) - 1:
        cut = last_ok + 1
        months_p = months_p[:cut]
        out_short = out_short[:cut]
        out_belly = out_belly[:cut]
        out_long  = out_long[:cut]
        totals    = totals[:cut]

    return {
        'months':   [str(m) for m in months_p],
        'short_bn': [round(v, 4) for v in out_short],
        'belly_bn': [round(v, 4) for v in out_belly],
        'long_bn':  [round(v, 4) for v in out_long],
        'total_bn': totals,
    }


def _file_mtime_str(path):
    try:
        if path.exists():
            return dt.datetime.fromtimestamp(path.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
    except Exception:
        pass
    return None


def _obligacje_hurtowe_latest_auction(yields_dir):
    """Read latest DataTransakcji from Obligacje_Hurtowe.xlsm / Operacje sheet.
    Returns (date_str, filename) or (None, filename) on failure.
    """
    path = yields_dir / 'Obligacje_Hurtowe.xlsm'
    if not path.exists():
        return None, None
    try:
        df = pd.read_excel(path, sheet_name='Operacje', usecols=['DataTransakcji'])
        df['DataTransakcji'] = pd.to_datetime(df['DataTransakcji'], errors='coerce')
        last = df['DataTransakcji'].max()
        if pd.isna(last):
            return None, path.name
        return last.strftime('%Y-%m-%d'), path.name
    except Exception:
        return None, path.name


def _kalkulator_last_updated(yields_dir):
    """Find a "Last updated on:" / "Ostatnia aktualizacja" / "Stan na" cell in
    kalkulatorodsetek.xlsm; parse Polish DD.MM.YYYY format if found as a string.
    Falls back to the workbook's `modified` core property (set by Excel when
    MoF saved the file), which is at the same level of trust.

    Returns (date_str, filename) or (None, filename) on failure.
    """
    import re as _re
    path = yields_dir / 'kalkulatorodsetek.xlsm'
    if not path.exists():
        return None, None
    try:
        import openpyxl as _opx
    except Exception:
        return None, path.name
    try:
        wb = _opx.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None, path.name

    # Phase 1: cell scan for the marker text on every sheet, first 50 rows.
    marker_re = _re.compile(r'last\s*updated|ostatnia\s*aktualizacja|stan\s*na',
                             _re.IGNORECASE)
    date_eu  = _re.compile(r'(\d{2})\.(\d{2})\.(\d{4})')
    for sn in wb.sheetnames:
        try:
            sh = wb[sn]
            for row in sh.iter_rows(max_row=50, max_col=20, values_only=True):
                # Look at every cell — if it matches the marker, try to find
                # an adjacent or trailing date.
                for i, v in enumerate(row):
                    if v is None: continue
                    sv = str(v)
                    if not marker_re.search(sv):
                        continue
                    # Date might be: (a) in the same cell appended after the colon
                    m = date_eu.search(sv)
                    if m:
                        dd, mm, yyyy = m.groups()
                        return f'{yyyy}-{mm}-{dd}', path.name
                    # (b) in the next non-empty cell to the right
                    for j in range(i + 1, len(row)):
                        nv = row[j]
                        if nv is None: continue
                        if isinstance(nv, dt.datetime):
                            return nv.strftime('%Y-%m-%d'), path.name
                        nm = date_eu.search(str(nv))
                        if nm:
                            dd, mm, yyyy = nm.groups()
                            return f'{yyyy}-{mm}-{dd}', path.name
                        break
        except Exception:
            continue

    # Phase 2: fall back to the workbook's modified property (set by Excel
    # when MoF saved the file — equivalent to "last touched by MoF").
    try:
        mod = wb.properties.modified
        if mod is not None:
            return mod.strftime('%Y-%m-%d'), path.name
    except Exception:
        pass
    return None, path.name


def _build_freshness(bases, nss, acm, brw):
    """Compile a `data lineage & freshness` block for the curves page.

    Reports the most recent observation timestamp for each input feeding the
    pipeline, plus the source filename (and lineage line for external sources)
    so advanced users can verify what's behind every chart.
    """
    fresh = {}

    # ---- BondSpot SA (scraped daily) ------------------------------------
    if bases is not None and not bases.empty:
        last_d = bases['date'].max()
        fresh['bondspot_fixings'] = {
            'last_obs': last_d.strftime('%Y-%m-%d'),
            'source':   'bases_merged_all.csv (scraped from BondSpot SA)',
            'detail':   f'{len(bases):,} bond-day rows since {bases["date"].min().strftime("%Y-%m-%d")}',
        }
        if 'turnover_value' in bases.columns:
            tv = pd.to_numeric(bases['turnover_value'], errors='coerce').fillna(0)
            nz = bases[tv > 0]
            if not nz.empty:
                fresh['bondspot_turnover'] = {
                    'last_obs': nz['date'].max().strftime('%Y-%m-%d'),
                    'source':   'bases_merged_all.csv · turnover_value column',
                    'detail':   'BondSpot SA venue daily turnover values',
                }

    # ---- Min-Fin PL Transakcje_po_seriach (monthly turnover) ------------
    # Single consolidated row: the actually-used xlsx filename + ref_month
    # of latest reported data.
    transakcje_files = sorted(YIELDS_DIR.glob('Transakcje_po_seriach*.xlsx'))
    latest_tx_name = transakcje_files[-1].name if transakcje_files else None
    last_ref_month = None
    if bases is not None and 'turnover_monthly_mln' in bases.columns and 'ref_month' in bases.columns:
        tm = pd.to_numeric(bases['turnover_monthly_mln'], errors='coerce').fillna(0)
        nz = bases[tm > 0]
        if not nz.empty:
            last_ref_month = str(nz['ref_month'].max())[:7]
    if latest_tx_name or last_ref_month:
        fresh['minfin_turnover'] = {
            'last_ref_month': last_ref_month or '',
            'source':         (latest_tx_name or 'Transakcje_po_seriach*.xlsx') + ' (Min-Fin PL)',
            'detail':         'Polish Ministry of Finance monthly per-series turnover, sheet "Obligacje(T-bonds)_outright". Publication lag up to ~2 months.',
        }

    # ---- Min-Fin PL Obligacje_Hurtowe (auctions + outstanding) ----------
    # Reports the latest DataTransakcji (most recent auction settled) so the
    # operator can verify no recent auction was missed.
    auction_date, ob_name = _obligacje_hurtowe_latest_auction(YIELDS_DIR)
    if ob_name:
        fresh['minfin_obligacje_hurtowe'] = {
            'last_auction':   auction_date or '',
            'source':         ob_name + ' (Min-Fin PL)',
            'detail':         'Polish Ministry of Finance auctions register and monthly outstanding amounts · latest DataTransakcji in Operacje sheet',
        }

    # ---- Min-Fin PL kalkulatorodsetek (coupon schedules) ----------------
    # Reports the "Last updated on:" cell (parsed as DD.MM.YYYY if present)
    # or, when absent, the workbook's `modified` core property as a fallback.
    upd_date, kalk_name = _kalkulator_last_updated(YIELDS_DIR)
    if kalk_name:
        fresh['minfin_coupon_schedules'] = {
            'last_updated':   upd_date or '',
            'source':         kalk_name + ' (Min-Fin PL)',
            'detail':         'Polish Ministry of Finance coupon-schedule reference workbook',
        }

    # ---- NBP MPC reference rate -----------------------------------------
    nbp_csv = YIELDS_DIR / 'nbp_policy_rates_live.csv'
    if nbp_csv.exists():
        try:
            nbp_df = pd.read_csv(nbp_csv)
            date_col = next((c for c in ('effective_date', 'date', 'effective') if c in nbp_df.columns), None)
            rate_col = next((c for c in ('rate_pct', 'rate', 'value', 'ref_rate') if c in nbp_df.columns), None)
            if date_col and not nbp_df.empty:
                nbp_df[date_col] = pd.to_datetime(nbp_df[date_col], errors='coerce')
                nbp_df = nbp_df.dropna(subset=[date_col]).sort_values(date_col)
                if not nbp_df.empty:
                    latest = nbp_df.iloc[-1]
                    rate = float(latest[rate_col]) if rate_col and pd.notna(latest.get(rate_col)) else None
                    fresh['nbp_reference_rate'] = {
                        'last_change': latest[date_col].strftime('%Y-%m-%d'),
                        'value':       (f'{rate:.2f}%' if rate is not None else None),
                        'source':      'nbp_policy_rates_live.csv (scraped from pl.wikipedia)',
                        'detail':      'NBP Monetary Policy Council reference-rate decisions',
                    }
        except Exception:
            pass

    # ---- FRB GSW (United States) ---------------------------------------
    # Reads the last data row of feds200628.csv (header narrative on
    # lines 1-9, real CSV header on line 10). Refreshed daily by
    # us_ea_autofetch.py.
    gsw_raw = YIELDS_DIR / 'feds200628.csv'
    if gsw_raw.exists():
        try:
            last_date = None
            with gsw_raw.open('r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    m = re.match(r'^(\d{4}-\d{2}-\d{2}),', line)
                    if m:
                        last_date = m.group(1)
            mtime = dt.datetime.fromtimestamp(gsw_raw.stat().st_mtime)
            fresh['fed_gsw_curve'] = {
                'last_obs':  last_date or '',
                'last_file': mtime.strftime('%Y-%m-%d %H:%M'),
                'source':    'feds200628.csv (Federal Reserve Board)',
                'detail':    'Gurkaynak-Sack-Wright daily Svensson '
                             'parameters for the US Treasury zero-coupon '
                             'yield curve. Auto-refreshed daily.',
            }
        except Exception:
            pass

    # ---- ECB AAA (euro area) -------------------------------------------
    # Reads the lex-latest ECB Data Portal_*.csv (cart-export wide
    # format with DATE on column 0). Refreshed daily by us_ea_autofetch.py.
    ecb_raws = sorted(YIELDS_DIR.glob('ECB Data Portal_*.csv'))
    if ecb_raws:
        ecb_raw = ecb_raws[-1]
        try:
            ecb_df = pd.read_csv(ecb_raw, usecols=[0])
            ecb_df.columns = ['DATE']
            ecb_df['DATE'] = pd.to_datetime(ecb_df['DATE'], errors='coerce')
            last_date = ecb_df['DATE'].dropna().max()
            mtime = dt.datetime.fromtimestamp(ecb_raw.stat().st_mtime)
            fresh['ecb_aaa_curve'] = {
                'last_obs':  (last_date.strftime('%Y-%m-%d')
                              if pd.notna(last_date) else ''),
                'last_file': mtime.strftime('%Y-%m-%d %H:%M'),
                'source':    ecb_raw.name + ' (ECB Data Portal)',
                'detail':    'European Central Bank daily Svensson '
                             'parameters for the AAA-rated euro-area '
                             'government bond yield curve. Auto-refreshed daily.',
            }
        except Exception:
            pass

    # ---- Treasury BondSpot Poland (TBSP) published index ----------------
    # Reads the last Date row of tbsp_d.csv (schema: Date,Open,High,Low,
    # Close,Volume). Refreshed daily by tbsp_autofetch.py (Step -1 of
    # update_daily.py, scrapes stooq.pl).
    tbsp_raw = YIELDS_DIR / 'tbsp_d.csv'
    if tbsp_raw.exists():
        try:
            tbsp_df = pd.read_csv(tbsp_raw, usecols=['Date'])
            tbsp_df['Date'] = pd.to_datetime(tbsp_df['Date'], errors='coerce')
            last_date = tbsp_df['Date'].dropna().max()
            mtime = dt.datetime.fromtimestamp(tbsp_raw.stat().st_mtime)
            fresh['tbsp_published_index'] = {
                'last_obs':  (last_date.strftime('%Y-%m-%d')
                              if pd.notna(last_date) else ''),
                'last_file': mtime.strftime('%Y-%m-%d %H:%M'),
                'source':    'tbsp_d.csv (scraped from stooq.pl, ticker tbsp.)',
                'detail':    'Treasury BondSpot Poland total-return index, '
                             'end-of-day; feeds the TR-decomposition tab and '
                             'the synthetic-TBSP validation. Auto-refreshed '
                             'daily by tbsp_autofetch.py (Step -1).',
            }
        except Exception:
            pass

    # ---- Pipeline outputs ----------------------------------------------
    if nss is not None and not nss.empty:
        fresh['nss_curves'] = {
            'last_obs': nss['tradedate'].max().strftime('%Y-%m-%d'),
            'source':   'nss_params_history.csv (computed in this pipeline)',
            'detail':   f'{len(nss):,} LW-NSS daily fits',
        }
    if acm is not None and not acm.empty:
        fresh['acm_term_premia'] = {
            'last_obs': acm['tradedate'].max().strftime('%Y-%m-%d'),
            'source':   'acm_term_premia_history.csv (computed in this pipeline)',
            'detail':   'Adrian-Crump-Moench affine TSM, weekly Friday closes',
        }
    if brw is not None and not brw.empty:
        fresh['brw_term_premia'] = {
            'last_obs': brw['tradedate'].max().strftime('%Y-%m-%d'),
            'source':   'acm_term_premia_brw_history.csv (computed in this pipeline)',
            'detail':   'Bauer-Rudebusch-Wu bias-corrected ACM',
        }

    return fresh


def _ffill_turnover(df):
    """Add a `turnover_filled_mln` column that ffills the last non-zero monthly
    turnover per ISIN. Returns the same df with the new column, or the df
    unchanged if it's None/empty or doesn't have the expected columns.

    The publication lag on KDPW / Min-Fin monthly turnover can be up to two
    months, so a snapshot taken before the latest report is released sees the
    most recent month as 0/NaN. We mask out 0/NaN and forward-fill within each
    ISIN, so the bond panel always shows the most recently *reported* turnover
    figure for each bond.
    """
    if df is None or df.empty:
        return df
    if 'turnover_monthly_mln' not in df.columns or 'ISIN' not in df.columns:
        return df
    df = df.sort_values(['ISIN', 'date']).reset_index(drop=True).copy()
    masked = df['turnover_monthly_mln'].where(
        pd.to_numeric(df['turnover_monthly_mln'], errors='coerce').fillna(0) > 0
    )
    df['turnover_filled_mln'] = masked.groupby(df['ISIN']).ffill()
    return df


# Polish sovereign-bond series naming convention: two-letter prefix (PS, DS, OK,
# WS, IZ, WZ, ...) + four-digit MMYY suffix. We approximate the maturity day as
# the 25th of MM — accurate to ±15 days, which translates to ≤0.04 yr on TTM.
_SERIES_PAT = re.compile(r'^([A-Z]{2,3})(\d{2})(\d{2})$')


def _parse_maturity_from_series(name):
    if not isinstance(name, str):
        return None
    m = _SERIES_PAT.match(name.strip())
    if not m:
        return None
    mm, yy = int(m.group(2)), int(m.group(3))
    year = 2000 + yy if yy < 80 else 1900 + yy
    if not (1 <= mm <= 12):
        return None
    try:
        return pd.Timestamp(year=year, month=mm, day=25)
    except Exception:
        return None


def _segment_from_ttm(ttm):
    if ttm is None:
        return ''
    if ttm <= 1.5:
        return '(0, 1.5]'
    if ttm <= 7.0:
        return '(1.5, 7]'
    return '(7, 30]'


def _venue_monthly_turnover(bases):
    """Per-(ISIN, month) BondSpot venue turnover in PLN mln.

    Sums the daily `turnover_value` column (already in PLN mln) over calendar
    months. This is the actual cash-market activity on the BondSpot venue,
    typically two orders of magnitude below the Min-Fin whole-market figure.
    """
    if bases is None or bases.empty or 'turnover_value' not in bases.columns:
        return {}
    b = bases[['ISIN', 'date', 'turnover_value']].copy()
    b['tv'] = pd.to_numeric(b['turnover_value'], errors='coerce').fillna(0.0)
    b['m'] = b['date'].dt.to_period('M')
    g = b.groupby(['ISIN', 'm'])['tv'].sum()
    return {k: float(v) for k, v in g.items()}


def _turnover_months(d, mf_months):
    """(prior calendar month, latest Min-Fin reported month <= prior) as str."""
    prior = d.to_period('M') - 1
    mf_m = None
    for m in reversed(mf_months):
        if m <= prior:
            mf_m = m
            break
    return str(prior), (str(mf_m) if mf_m is not None else None)


def _turnover_cols(isin, name, d, venue_tov, mf_long, mf_months):
    """Three display turnover figures for one bond on snapshot date d.

    Returns (tov_bs, tov_mf, tov_share):
      tov_bs    BondSpot venue turnover in the prior calendar month, PLN mln
      tov_mf    Min-Fin whole-market OUTRIGHT turnover for the series in the
                latest reported month at or before the prior month, PLN bn
      tov_share BondSpot share of that outright month, in percent, computed
                venue-vs-Min-Fin on the SAME month so the ratio is honest
                even when the Min-Fin publication lags the venue data
    """
    prior = d.to_period('M') - 1
    bs = venue_tov.get((isin, prior))
    mf_m = None
    for m in reversed(mf_months):
        if m <= prior:
            mf_m = m
            break
    mf_bn = share = None
    if mf_m is not None:
        v = mf_long.get((name, mf_m))
        if v is not None and v > 0:
            mf_bn = v
            bs_at_mf = venue_tov.get((isin, mf_m))
            if bs_at_mf is not None:
                share = 100.0 * bs_at_mf / (v * 1000.0)
    return (round(bs, 1) if bs is not None else None,
            round(mf_bn, 2) if mf_bn is not None else None,
            round(share, 2) if share is not None else None)


def _bond_panel(bond, d, venue_tov, mf_long, mf_months):
    if bond is None or bond.empty:
        return []
    sub = bond[bond['date'] == d]
    if sub.empty:
        return []
    rows = []
    for _, b in sub.iterrows():
        out_mln = b.get('outstanding_mln')
        isin = str(b.get('ISIN', ''))
        name = str(b.get('Nazwa', '')).strip()
        tov_bs, tov_mf, tov_share = _turnover_cols(
            isin, name, d, venue_tov, mf_long, mf_months)
        rows.append({
            'isin':      isin,
            'name':      name,
            'ttm':       float(b['ttm']) if pd.notna(b.get('ttm')) else None,
            'ytm':       float(b['rent_fix_pct']) if pd.notna(b.get('rent_fix_pct')) else None,
            'out':       round(float(out_mln) / 1000.0, 2) if pd.notna(out_mln) else None,  # PLN bn
            'tov_bs':    tov_bs,     # PLN mln, BondSpot venue, prior month
            'tov_mf':    tov_mf,     # PLN bn, Min-Fin outright, latest reported month
            'tov_share': tov_share,  # percent, venue / outright, same month
            'segment':   str(b.get('segment', '')),
        })
    return rows


def _bond_panel_from_bases(bases, d, venue_tov, mf_long, mf_months):
    """Fallback bond panel for snapshots newer than the last liquidity rebuild.

    Reads the daily BondSpot+weights panel and computes TTM from the series-name
    suffix (PS/DS/OK + MMYY). Segment is derived from TTM with the same bin
    boundaries as the liquidity_bond_day_full schema.
    """
    if bases is None or bases.empty:
        return []
    sub = bases[bases['date'] == d]
    if sub.empty:
        return []
    rows = []
    for _, b in sub.iterrows():
        name = str(b.get('Nazwa', '')).strip()
        mat = _parse_maturity_from_series(name)
        if mat is None or mat <= d:
            continue
        ttm = max(0.0, (mat - d).days / 365.25)
        out_mln = b.get('outstanding_mln')
        ytm = b.get('rent_fix_pct')
        if pd.isna(ytm):
            continue  # no fixing today, skip
        isin = str(b.get('ISIN', ''))
        tov_bs, tov_mf, tov_share = _turnover_cols(
            isin, name, d, venue_tov, mf_long, mf_months)
        rows.append({
            'isin':      isin,
            'name':      name,
            'ttm':       round(ttm, 4),
            'ytm':       float(ytm),
            'out':       round(float(out_mln) / 1000.0, 2) if pd.notna(out_mln) else None,
            'tov_bs':    tov_bs,
            'tov_mf':    tov_mf,
            'tov_share': tov_share,
            'segment':   _segment_from_ttm(ttm),
        })
    rows.sort(key=lambda r: r['ttm'])
    return rows


def _bond_hist(bond, isins):
    if bond is None or bond.empty or not isins:
        return {}
    out = {}
    for isin in isins:
        sub = bond[bond['ISIN'] == isin].copy()
        if sub.empty:
            continue
        sub['week'] = sub['date'].dt.to_period('W-FRI')
        sub_w = sub.sort_values('date').groupby('week').last().reset_index(drop=True)
        sub_w = sub_w.dropna(subset=['rent_fix_pct', 'ttm'])
        if sub_w.empty:
            continue
        out[isin] = {
            'name':  str(sub_w['Nazwa'].iloc[-1]),
            'dates': sub_w['date'].dt.strftime('%Y-%m-%d').tolist(),
            'ytm':   [round(float(v), 3) for v in sub_w['rent_fix_pct']],
            'ttm':   [round(float(v), 4) for v in sub_w['ttm']],
        }
    return out


def _eh_fb1_xm_heatmap():
    """Cross-market FB1 β on a common (h, n) grid for PL, US, EA — Figure 2 of EH paper."""
    out = {}
    for tag, fn in (('pl', 'eh_fb1_pl_xm.csv'),
                    ('us', 'eh_fb1_us_xm.csv'),
                    ('ea', 'eh_fb1_ea_xm.csv')):
        df = _try_read(fn)
        if df is None:
            return None
        # Build (h_m × n_m) pivot of beta
        pv = df.pivot_table(index='h_m', columns='n_m', values='beta', aggfunc='first')
        out[tag] = {
            'h_m':    [int(x) for x in pv.index.tolist()],
            'n_m':    [int(x) for x in pv.columns.tolist()],
            'matrix': [[None if pd.isna(v) else round(float(v), 3) for v in row]
                       for row in pv.values.tolist()],
        }
    return out


def _eh_macro_r2_by_h():
    """Median R² across (h, n) pairs by horizon for the three nested macro-spanning specs."""
    df = _try_read('eh_macro_spanning_results.csv')
    if df is None:
        return None
    df = df[df['applicable'] == True].copy()
    g = df.groupby('h_m').agg(
        r2_fb=('r2_fb_only', 'median'),
        r2_macro=('r2_macro_only', 'median'),
        r2_joint=('r2_joint', 'median'),
        n_pairs=('n_obs', 'count'),
    ).reset_index()
    return {
        'h_m':       [int(x) for x in g['h_m']],
        'r2_fb':     [round(float(v), 4) for v in g['r2_fb']],
        'r2_macro':  [round(float(v), 4) for v in g['r2_macro']],
        'r2_joint':  [round(float(v), 4) for v in g['r2_joint']],
        'n_pairs':   [int(x) for x in g['n_pairs']],
    }


def _fit_xcountry_acm(zero_csv, cadence='monthly'):
    """Fit ACM 5-factor model on a daily/monthly zero panel, return DataFrame with
    term-premia at 1y, 2y, 3y, 5y, 7y, 10y. Reuses acm_tp.estimate_acm
    and acm_tp.project_daily so the methodology matches the PL fit.

    cadence='monthly' (default): one row per calendar month (last daily fit).
    cadence='daily': all daily fits (used by xcountry_tp_d build)."""
    import sys as _sys
    if str(YIELDS_DIR) not in _sys.path:
        _sys.path.insert(0, str(YIELDS_DIR))
    import acm_tp
    from scipy.interpolate import CubicSpline

    df = pd.read_csv(zero_csv, parse_dates=['date']).sort_values('date').reset_index(drop=True)
    cols = ['y_3m','y_6m','y_1y','y_2y','y_3y','y_4y','y_5y','y_6y','y_7y','y_8y','y_9y','y_10y']
    tenors = np.array([0.25, 0.5, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    Y = df[cols].values  # cc-yields as decimal fractions
    TTM = np.linspace(1/12, 10, 120)

    Y_full = np.full((len(df), len(TTM)), np.nan)
    for i in range(len(df)):
        row = Y[i]
        if np.isnan(row).any():
            continue
        cs = CubicSpline(tenors, row)
        Y_full[i] = cs(TTM)
    valid = ~np.isnan(Y_full).any(axis=1)
    df = df.loc[valid].reset_index(drop=True)
    Y_full = Y_full[valid]

    daily_df = pd.DataFrame(Y_full, index=df['date'])
    monthly_df = daily_df.resample('MS').first().dropna(how='all')
    Y_monthly = monthly_df.values

    model = acm_tp.estimate_acm(Y_monthly, TTM, K=5, min_pca_tenor_months=3,
                                dates_monthly=monthly_df.index)
    daily_out = acm_tp.project_daily(Y_full, model)
    tp = daily_out['term_premia_cc']  # cc, decimal

    horizons = {'1y': 12, '2y': 24, '3y': 36, '5y': 60, '7y': 84, '10y': 120}
    out = pd.DataFrame({'date': df['date']})
    for k, m in horizons.items():
        out[f'tp_{k}_bp'] = tp[:, m - 1] * 1e4
    if cadence == 'daily':
        daily = out.copy()
        daily['date'] = daily['date'].dt.strftime('%Y-%m-%d')
        return daily[['date'] + [c for c in daily.columns if c.startswith('tp_')]].copy()
    out['m'] = out['date'].dt.to_period('M')
    monthly = out.groupby('m').last().reset_index()
    monthly['date'] = monthly['m'].astype(str)
    return monthly[['date'] + [c for c in monthly.columns if c.startswith('tp_')]].copy()


def _build_xcountry_tp(us_csv, ea_csv, brw):
    """Return monthly cross-country term-premium panel with PL BRW + US ACM + EA ACM."""
    out = {'dates': []}
    horizons = ['1y', '2y', '3y', '5y', '7y', '10y']
    for tag in ('pl', 'us', 'ea'):
        for h in horizons:
            out[f'{tag}_tp_{h}_bp'] = []

    # PL BRW monthly resample (last obs of month from weekly BRW history)
    pl = brw.copy()
    pl['m'] = pl['tradedate'].dt.to_period('M')
    pl_m = pl.sort_values('tradedate').groupby('m').last().reset_index()
    pl_m['date'] = pl_m['m'].astype(str)

    # US / EA ACM fits
    us_m = _fit_xcountry_acm(us_csv) if us_csv.exists() else None
    ea_m = _fit_xcountry_acm(ea_csv) if ea_csv.exists() else None
    if us_m is not None: print(f'  US ACM fit: {len(us_m)} months')
    if ea_m is not None: print(f'  EA ACM fit: {len(ea_m)} months')

    months = sorted(set(pl_m['date'])
                    & set(us_m['date'] if us_m is not None else [])
                    & set(ea_m['date'] if ea_m is not None else []))
    for d in months:
        out['dates'].append(d)
        plr = pl_m[pl_m['date'] == d].iloc[0]
        usr = us_m[us_m['date'] == d].iloc[0]  if us_m is not None else None
        ear = ea_m[ea_m['date'] == d].iloc[0]  if ea_m is not None else None
        for h in horizons:
            out[f'pl_tp_{h}_bp'].append(round(float(plr[f'tp_bc_{h}_bp']), 1) if pd.notna(plr[f'tp_bc_{h}_bp']) else None)
            out[f'us_tp_{h}_bp'].append(round(float(usr[f'tp_{h}_bp']), 1) if (usr is not None and pd.notna(usr[f'tp_{h}_bp'])) else None)
            out[f'ea_tp_{h}_bp'].append(round(float(ear[f'tp_{h}_bp']), 1) if (ear is not None and pd.notna(ear[f'tp_{h}_bp'])) else None)
    return out


def _build_xcountry_tp_daily(us_csv, ea_csv, brw):
    """Daily cross-country term-premium panel: PL BRW + US ACM + EA ACM on the
    common daily grid. Mirrors `_build_xcountry_tp` but at daily cadence so the
    cross-country chart on the term-premia page can be fresh."""
    out = {'dates': []}
    horizons = ['1y', '2y', '3y', '5y', '7y', '10y']
    for tag in ('pl', 'us', 'ea'):
        for h in horizons:
            out[f'{tag}_tp_{h}_bp'] = []

    pl = brw.sort_values('tradedate').copy()
    pl['date'] = pl['tradedate'].dt.strftime('%Y-%m-%d')
    pl_keep = ['date'] + [f'tp_bc_{h}_bp' for h in horizons]
    pl = pl[pl_keep].drop_duplicates(subset='date').set_index('date')
    pl = pl.rename(columns={f'tp_bc_{h}_bp': f'pl_tp_{h}_bp' for h in horizons})

    us_d = _fit_xcountry_acm(us_csv, cadence='daily') if us_csv.exists() else None
    ea_d = _fit_xcountry_acm(ea_csv, cadence='daily') if ea_csv.exists() else None
    if us_d is not None:
        us_d = us_d.drop_duplicates(subset='date').set_index('date')
        us_d = us_d.rename(columns={f'tp_{h}_bp': f'us_tp_{h}_bp' for h in horizons})
        print(f'  US ACM daily fit: {len(us_d)} days')
    if ea_d is not None:
        ea_d = ea_d.drop_duplicates(subset='date').set_index('date')
        ea_d = ea_d.rename(columns={f'tp_{h}_bp': f'ea_tp_{h}_bp' for h in horizons})
        print(f'  EA ACM daily fit: {len(ea_d)} days')

    joined = pl.copy()
    if us_d is not None:
        joined = joined.join(us_d, how='inner')
    if ea_d is not None:
        joined = joined.join(ea_d, how='inner')
    joined = joined.sort_index()

    out['dates'] = list(joined.index)
    for tag in ('pl', 'us', 'ea'):
        for h in horizons:
            col = f'{tag}_tp_{h}_bp'
            if col in joined.columns:
                out[col] = [round(float(v), 1) if pd.notna(v) else None for v in joined[col]]
            else:
                out[col] = [None] * len(joined)
    return out


def _tp_xcountry_corr(xtp, other='us'):
    """6×6 correlation matrix between PL BRW TPs and {US|EA} ACM TPs at 1/2/3/5/7/10y."""
    horizons = ['1y', '2y', '3y', '5y', '7y', '10y']
    if not xtp or not xtp.get('dates'):
        return None
    df_pl = pd.DataFrame({h: xtp[f'pl_tp_{h}_bp'] for h in horizons})
    df_x  = pd.DataFrame({h: xtp[f'{other}_tp_{h}_bp'] for h in horizons})
    df = pd.concat([df_pl.add_suffix('_pl'), df_x.add_suffix(f'_{other}')], axis=1).dropna()
    if df.empty:
        return None
    M = np.zeros((len(horizons), len(horizons)))
    for i, hi in enumerate(horizons):
        for j, hj in enumerate(horizons):
            M[i, j] = df[f'{hi}_pl'].corr(df[f'{hj}_{other}'])
    return {
        'pl_labels': [f'PL BRW {h}' for h in horizons],
        'x_labels':  [f'{other.upper()} ACM {h}' for h in horizons],
        'matrix':    np.round(M, 3).tolist(),
    }


# ------------------------------ entry ------------------------------ #

def main(argv=None):
    global YIELDS_DIR, DIST, DATA_DIR
    ap = argparse.ArgumentParser(description='Build yieldcartography.com yields.json')
    ap.add_argument('--out-dir', type=str, default=None,
                    help='Destination dist/ directory. yields.json lands in <out-dir>/data/. '
                         'Default: ./dist (relative to this script).')
    ap.add_argument('--yields-dir', type=str, default=None,
                    help='Override the YIELDS folder where the CSV inputs live. '
                         f'Default: {YIELDS_DIR}')
    args = ap.parse_args(argv)

    if args.yields_dir:
        YIELDS_DIR = Path(args.yields_dir).expanduser().resolve()
    if args.out_dir:
        DIST = Path(args.out_dir).expanduser().resolve()
        DATA_DIR = DIST / 'data'

    print('building yieldcartography.com...')
    print('reading data from', YIELDS_DIR)
    print('writing dist to ', DIST)
    data = build()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    out_path = DATA_DIR / 'yields.json'
    out_path.write_text(json.dumps(data, separators=(',', ':')))
    try:
        rel = out_path.relative_to(SITE_DIR)
    except ValueError:
        rel = out_path  # absolute path is fine when writing to an external repo
    print(f'  wrote {rel} ({out_path.stat().st_size/1024:.1f} KB)')
    print(f'  meta: {data["meta"]}')

    # Regenerate the Oracle tab from the same YIELDS CSVs so it tracks every
    # data refresh alongside yields.json. Non-fatal: a failure here must not
    # block the rest of the site build.
    try:
        import sys as _sys
        if str(YIELDS_DIR) not in _sys.path:
            _sys.path.insert(0, str(YIELDS_DIR))
        import oracle_build
        oracle_out = DIST / 'oracle' / 'index.html'
        oracle_build.write_page(oracle_out)
        print(f'  wrote {oracle_out.relative_to(DIST.parent) if DIST.parent in oracle_out.parents else oracle_out} (Oracle tab)')
    except Exception as e:
        print(f'  WARNING: Oracle tab not regenerated ({type(e).__name__}: {e})')

    print('done.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
